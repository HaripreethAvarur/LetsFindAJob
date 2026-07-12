"""job_tracker.xlsx — the human-facing view, regenerated from SQLite each run.

The workbook is disposable: if it is corrupted by manual editing while the
pipeline runs, delete it — the next run rebuilds it from the database with
no data loss. Manual tracking edits belong in the application_status table
(or its CSV), never in this file.
"""
from __future__ import annotations

import sqlite3
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

VERDICT_FILLS = {
    "APPLY": PatternFill("solid", fgColor="C6EFCE"),   # green
    "MAYBE": PatternFill("solid", fgColor="FFEB9C"),   # yellow
    "SKIP": PatternFill("solid", fgColor="D9D9D9"),    # gray
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")

LISTING_QUERY = """
SELECT j.job_id, j.company, j.title, j.location, j.posted_date, j.source_platform,
       j.url, j.first_seen, j.last_change, j.is_stale, j.duplicate_of,
       e.relevance_score, e.verdict, e.skill_gap, e.reasoning,
       e.recruiter_contact, e.contact_source,
       a.status, a.follow_up_date, a.notes, a.resume_path, a.cover_letter_path
FROM jobs_master j
LEFT JOIN ai_evaluations e USING (job_id)
LEFT JOIN application_status a USING (job_id)
"""

HEADERS = [
    "Job ID", "Company", "Title", "Location", "Posted", "Source", "URL",
    "First Seen", "Change", "Stale", "Duplicate Of", "Score", "Verdict",
    "Skill Gap", "Reasoning", "Recruiter Contact", "Contact Source",
    "Status", "Follow-up", "Notes", "Resume", "Cover Letter",
]


def _style_sheet(ws: Worksheet, verdict_col: int) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    # Conditional fill by verdict, applied to the whole row.
    for row in ws.iter_rows(min_row=2):
        verdict = row[verdict_col - 1].value
        fill = VERDICT_FILLS.get(str(verdict or "").upper())
        if fill:
            for cell in row:
                cell.fill = fill
    # Auto-size columns (capped so descriptions don't explode the sheet).
    for idx, column in enumerate(ws.columns, start=1):
        width = max((len(str(c.value)) for c in column if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 9), 55)


def _write_rows(ws: Worksheet, rows: Iterable[sqlite3.Row]) -> None:
    ws.append(HEADERS)
    for r in rows:
        ws.append([
            r["job_id"], r["company"], r["title"], r["location"], r["posted_date"],
            r["source_platform"], r["url"], r["first_seen"], r["last_change"],
            "yes" if r["is_stale"] else "", r["duplicate_of"], r["relevance_score"],
            r["verdict"], r["skill_gap"], r["reasoning"], r["recruiter_contact"],
            r["contact_source"], r["status"], r["follow_up_date"], r["notes"],
            r["resume_path"], r["cover_letter_path"],
        ])
    _style_sheet(ws, verdict_col=HEADERS.index("Verdict") + 1)


def write_tracker(
    conn: sqlite3.Connection,
    out_path: Path,
    run_started_iso: str,
    run_summary: dict[str, Any] | None = None,
) -> Path:
    wb = Workbook()

    # --- New This Run -------------------------------------------------------
    ws_new = wb.active
    assert ws_new is not None
    ws_new.title = "New This Run"
    new_rows = conn.execute(
        LISTING_QUERY
        + " WHERE j.last_change IN ('new','changed') AND j.last_change_at >= ?"
        " ORDER BY e.relevance_score DESC NULLS LAST, j.first_seen DESC",
        (run_started_iso,),
    ).fetchall()
    _write_rows(ws_new, new_rows)

    # --- All Listings -------------------------------------------------------
    ws_all = wb.create_sheet("All Listings")
    all_rows = conn.execute(
        LISTING_QUERY + " ORDER BY e.relevance_score DESC NULLS LAST, j.first_seen DESC"
    ).fetchall()
    _write_rows(ws_all, all_rows)

    # --- Stats --------------------------------------------------------------
    ws_stats = wb.create_sheet("Stats")
    _write_stats(ws_stats, conn, run_summary or {})

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _write_stats(ws: Worksheet, conn: sqlite3.Connection, run_summary: dict[str, Any]) -> None:
    from . import db as dbmod

    bold = Font(bold=True)
    row = 1

    def title(text: str) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = bold
        row += 1

    def kv(key: str, value: Any) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=str(value))
        row += 1

    # Follow-up reminders surface at the very top, per spec.
    title("⚠ FOLLOW-UPS OVERDUE (status=Applied, follow_up_date passed)")
    overdue = dbmod.overdue_followups(conn)
    if overdue:
        for r in overdue:
            kv(f"{r['company']} — {r['title']}", f"follow-up was due {r['follow_up_date']} · {r['url']}")
    else:
        kv("(none)", "")
    row += 1

    title("APPLICATION FUNNEL")
    funnel = dbmod.funnel_stats(conn)
    for stage in ("New", "Applied", "Response", "Interview", "Offer", "Rejected"):
        kv(stage, funnel.get(stage, 0))
    for other, n in sorted(funnel.items()):
        if other not in ("New", "Applied", "Response", "Interview", "Offer", "Rejected"):
            kv(other, n)
    row += 1

    title("VERDICT BREAKDOWN")
    for r in conn.execute(
        "SELECT verdict, COUNT(*) n FROM ai_evaluations GROUP BY verdict ORDER BY n DESC"
    ):
        kv(r["verdict"] or "(unscored)", r["n"])
    row += 1

    title("LISTINGS BY SOURCE")
    for r in conn.execute(
        "SELECT source_platform, COUNT(*) n FROM jobs_master GROUP BY source_platform ORDER BY n DESC"
    ):
        kv(r["source_platform"], r["n"])
    row += 1

    if run_summary:
        title("LAST RUN")
        for key, value in run_summary.items():
            kv(key, value)

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 70
    ws.freeze_panes = "A2"
