from openpyxl import load_workbook

from core import db
from core.dedup import process_record
from core.excel_writer import write_tracker
from core.models import Evaluation, JobRecord


def seed(conn):
    for i, verdict in enumerate(["APPLY", "MAYBE", "SKIP"]):
        rec = JobRecord(
            job_id=f"greenhouse:acme:{i}",
            company="Acme",
            title=f"Role {i}",
            location="Hyderabad",
            source_platform="greenhouse",
            raw_description=f"desc {i}",
        )
        process_record(conn, rec)
        db.ensure_application_row(conn, rec.job_id)
        db.save_evaluation(
            conn,
            Evaluation(
                job_id=rec.job_id, content_hash=rec.content_hash,
                pass1_requirements={"hard_requirements": []},
                pass2_verification={"checks": []},
                relevance_score=90 - i * 30, skill_gap="none",
                verdict=verdict, reasoning="r",
            ),
        )


def test_tracker_workbook(tmp_path):
    conn = db.connect(tmp_path / "t.db")
    seed(conn)
    out = tmp_path / "job_tracker.xlsx"
    write_tracker(conn, out, run_started_iso="2000-01-01T00:00:00+00:00",
                  run_summary={"llm_calls_made": 6})
    wb = load_workbook(out)

    assert wb.sheetnames == ["New This Run", "All Listings", "Stats"]

    ws = wb["All Listings"]
    assert ws.freeze_panes == "A2"
    header = [c.value for c in ws[1]]
    assert header[0] == "Job ID" and "Verdict" in header

    # 3 seeded rows + header
    assert ws.max_row == 4

    # Conditional formatting: APPLY row is green, SKIP row gray.
    verdict_col = header.index("Verdict") + 1
    fills = {}
    for row in ws.iter_rows(min_row=2):
        fills[row[verdict_col - 1].value] = row[0].fill.fgColor.rgb
    assert fills["APPLY"].endswith("C6EFCE")
    assert fills["SKIP"].endswith("D9D9D9")

    # Everything seeded this run counts as new.
    assert wb["New This Run"].max_row == 4

    stats_text = " ".join(
        str(c.value) for row in wb["Stats"].iter_rows() for c in row if c.value
    )
    assert "FOLLOW-UPS" in stats_text and "APPLICATION FUNNEL" in stats_text


def test_workbook_is_regenerable(tmp_path):
    """The xlsx is disposable: deleting it and re-running rebuilds identically
    from SQLite (no state lives in the file)."""
    conn = db.connect(tmp_path / "t.db")
    seed(conn)
    out = tmp_path / "job_tracker.xlsx"
    write_tracker(conn, out, "2000-01-01T00:00:00+00:00")
    out.unlink()
    write_tracker(conn, out, "2000-01-01T00:00:00+00:00")
    assert load_workbook(out)["All Listings"].max_row == 4
